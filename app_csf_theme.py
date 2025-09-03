
import io
import datetime as dt
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from zipfile import ZipFile, ZIP_DEFLATED

# ----------------------------
# Page config
# ----------------------------
st.set_page_config(
    page_title="FLN Monitoring Report Summary",
    page_icon=None,
    layout="wide"
)

# ----------------------------
# CSF-like color theme (blue/yellow)
# ----------------------------
PRIMARY = "#0B3D91"   # deep blue
ACCENT  = "#F7C948"   # warm yellow
BG      = "#F7F9FC"   # very light background
TEXT    = "#0F172A"   # slate-900
MUTED   = "#6B7280"   # slate-500
CARD_BG = "#FFFFFF"

# Global style
st.markdown(f"""
<style>
:root {{
  --primary: {PRIMARY};
  --accent: {ACCENT};
  --bg: {BG};
  --text: {TEXT};
  --muted: {MUTED};
  --card: {CARD_BG};
}}

html, body, .block-container {{
  background: var(--bg);
}}

#MainMenu, footer {{visibility: hidden;}}

.app-wrap {{
  padding-top: 8px;
}}

.app-title {{
    font-size: 34px;
    font-weight: 800;
    color: var(--text);
    text-align: center;
    letter-spacing: .2px;
}}
.app-subtitle {{
    font-size: 16px;
    text-align: center;
    color: var(--muted);
    margin-top: 4px;
}}

/* Divider underline in accent */
.underline {{
  width: 120px;
  height: 4px;
  background: var(--accent);
  margin: 10px auto 22px auto;
  border-radius: 999px;
}}

/* Section headings with blue left bar */
h3 {{
  color: var(--text);
  border-left: 6px solid var(--primary);
  padding-left: 12px;
  margin-top: 26px;
}}

/* Card containers */
.st-emotion-cache-16idsys, .st-emotion-cache-0, .st-emotion-cache-1r6slb0 {{
  background: var(--card);
  border: 1px solid #e5e7eb;
  border-radius: 14px;
  box-shadow: 0 6px 18px rgba(11,61,145,0.06);
  padding: 0.75rem 1rem 1rem 1rem;
}}

/* Inputs & buttons */
.stTextInput > div > div > input,
.stFileUploader, .stSelectbox, .stMultiSelect {{
  border-radius: 10px !important;
}}

.stButton > button, .stDownloadButton > button {{
  background: var(--primary);
  color: white;
  border-radius: 10px;
  padding: 0.6rem 1.1rem;
  font-weight: 600;
  border: 2px solid var(--primary);
}}

.stButton > button:hover, .stDownloadButton > button:hover {{
  background: #082e70;
  border-color: #082e70;
}}

.badge {{
  display: inline-block;
  background: var(--accent);
  color: #111827;
  font-weight: 700;
  padding: 4px 10px;
  border-radius: 999px;
  font-size: 12px;
  margin-left: 8px;
}}
</style>
""", unsafe_allow_html=True)

st.markdown("<div class='app-wrap'>", unsafe_allow_html=True)
st.markdown("<div class='app-title'>FLN Monitoring Report Summary</div>", unsafe_allow_html=True)
st.markdown("<div class='app-subtitle'>Generate FLN Block &amp; Cluster Monitoring Visit summary</div>", unsafe_allow_html=True)
st.markdown("<div class='underline'></div>", unsafe_allow_html=True)

# ----------------------------
# Helper functions (same logic as previous)
# ----------------------------
def detect_columns(df: pd.DataFrame):
    cols = list(df.columns)

    def guess(patterns):
        for p in patterns:
            for c in cols:
                try:
                    if pd.Series([str(c)]).str.contains(p, case=False, regex=True).iloc[0]:
                        return c
                except Exception:
                    pass
        return None

    district = guess([r"district", r"जिला"])
    block    = guess([r"block", r"ब्लॉक|खंड"])
    cluster  = guess([r"cluster|sankul|jsk|jan\\s*shiksha", r"संकुल|जन शिक्ष"])
    role     = guess([r"role", r"पद"])
    target   = guess([r"target|लक्ष्य"])
    done     = guess([r"completed|done|स्थिति|visit.*done|actual|completion"])

    return district, block, cluster, role, target, done

def build_table(df, report_type, district_col, block_col, cluster_col, role_col, target_col, done_col, district_value):
    base = df[df[district_col].astype(str).str.contains(str(district_value), case=False, na=False)].copy()
    if report_type == "CAC":
        if role_col and base[role_col].notna().any():
            mask = base[role_col].astype(str).str.upper().str.contains("CAC")
            if mask.any():
                base = base[mask]
        group_key = block_col
    else:  # BRC+BAC
        if role_col:
            mask = base[role_col].astype(str).str.upper().isin(["BRC", "BRCC", "BAC"])
            base = base[mask]
        group_key = block_col

    agg = base.groupby(group_key)[[target_col, done_col]].sum().reset_index()
    agg.rename(columns={group_key: "Block Name", target_col: "Targeted Visits", done_col: "Completed"}, inplace=True)
    agg["Monitoring %"] = np.where(agg["Targeted Visits"] > 0, (agg["Completed"]/agg["Targeted Visits"]) * 100, np.nan).round(0)
    agg = agg.sort_values("Monitoring %", ascending=False)

    grand = pd.DataFrame({
        "Block Name": ["Grand Total"],
        "Targeted Visits": [agg["Targeted Visits"].sum()],
        "Completed": [agg["Completed"].sum()],
        "Monitoring %": [round((agg["Completed"].sum()/agg["Targeted Visits"].sum())*100, 0) if agg["Targeted Visits"].sum() else np.nan]
    })
    final_tbl = pd.concat([agg, grand], ignore_index=True)
    return final_tbl

def style_and_save_excel(df_out, filename_prefix, district):
    from openpyxl import load_workbook
    with pd.ExcelWriter(io.BytesIO(), engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.book["Summary"]

        widths = [28, 18, 18, 16]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+i)].width = w

        header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill

        for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
            for cell in row:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center")

        for r in range(2, ws.max_row+1):
            cell = ws.cell(row=r, column=4)
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value/100.0

        last = ws.max_row
        for cell in ws[last]:
            cell.font = Font(bold=True)

        rule = ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                              mid_type='num', mid_value=0.65, mid_color='FFEB84',
                              end_type='num', end_value=1, end_color='63BE7B')
        ws.conditional_formatting.add(f"D2:D{last}", rule)

        writer.close()
        data = writer._handles.handle.getvalue()

    today = dt.datetime.now().strftime("%d-%m-%Y")
    filename = f"{filename_prefix}_{district}_{today}.xlsx"
    return data, filename

# ----------------------------
# UI Controls
# ----------------------------
st.header("Upload CSV")
file = st.file_uploader("Upload the raw state export (CSV)", type=["csv"])

c1, c2 = st.columns([1.2,1])
with c1:
    report_type = st.radio(
        "Report type",
        ["CAC (Cluster / CACs)", "BRC+BAC (Block Mentors)"],
        horizontal=True
    )
report_key = "CAC" if report_type.startswith("CAC") else "BRCBAC"
st.markdown(f"<span class='badge'>{'Cluster' if report_key=='CAC' else 'Block'}</span>", unsafe_allow_html=True)

if file is None:
    st.info("Upload a CSV to begin.")
    st.stop()

df = pd.read_csv(file, encoding_errors="ignore")
district_col, block_col, cluster_col, role_col, target_col, done_col = detect_columns(df)

needed = [district_col, block_col, target_col, done_col]
if not all(needed):
    st.error("Could not detect required columns. Ensure the CSV has at least District, Block, Targeted Visits, and Completed columns.")
    st.write("Detected:", dict(district=district_col, block=block_col, role=role_col, target=target_col, completed=done_col))
    st.stop()

st.markdown("**Detected columns**")
st.caption(f"District → `{district_col}` • Block → `{block_col}` • Target → `{target_col}` • Completed → `{done_col}` • Role → `{role_col or 'N/A'}`")

districts = sorted(df[district_col].dropna().unique().tolist())
mode = st.radio("Generate for:", ["Selected districts", "All districts"], horizontal=True)
default_sel = [d for d in districts if 'jabalpur' in str(d).lower()]
if mode == "Selected districts":
    sel = st.multiselect("Choose district(s)", districts, default=default_sel or districts[:1])
else:
    sel = districts

# ----------------------------
# Build Reports
# ----------------------------
build = st.button("⚙️ Build Reports", type="primary")
if build:
    outputs = []
    progress = st.progress(0, text="Starting...")
    total = len(sel)

    for i, dist in enumerate(sel, start=1):
        table = build_table(df, report_key, district_col, block_col, cluster_col, role_col, target_col, done_col, dist)
        prefix = "CAC_Monitoring" if report_key == "CAC" else "BRC_BAC_Monitoring"
        excel_bytes, xlsx_name = style_and_save_excel(table, prefix, dist)
        outputs.append((dist, table, excel_bytes, xlsx_name))
        progress.progress(i/total, text=f"Processed {i}/{total}: {dist}")

    progress.empty()

    st.header("Preview")
    st.dataframe(outputs[0][1], use_container_width=True)

    st.header("Downloads")
    for dist, table, excel_bytes, xlsx_name in outputs:
        st.download_button(
            label=f"⬇️ Download Excel – {dist}",
            data=excel_bytes,
            file_name=xlsx_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    if len(outputs) > 1:
        mem_zip = io.BytesIO()
        with ZipFile(mem_zip, mode="w", compression=ZIP_DEFLATED) as zf:
            for _, _, excel_bytes, xlsx_name in outputs:
                zf.writestr(xlsx_name, excel_bytes)
        st.download_button(
            label="⬇️ Download ALL as ZIP",
            data=mem_zip.getvalue(),
            file_name=f"FLN_{report_key}_Reports_{dt.datetime.now().strftime('%d-%m-%Y')}.zip",
            mime="application/zip"
        )

    st.success(f"Generated {len(outputs)} report(s).")

st.markdown("</div>", unsafe_allow_html=True)
