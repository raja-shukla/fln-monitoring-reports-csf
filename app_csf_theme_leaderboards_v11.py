
import io
import datetime as dt
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from zipfile import ZipFile, ZIP_DEFLATED

st.set_page_config(page_title="FLN Monitoring Report Summary", page_icon="📊", layout="wide")

# =====================
# CSF Theme & UI Shell
# =====================
st.markdown("""
<style>
:root {
  --primary:#0B3D91;   /* deep CSF blue */
  --accent:#F7C948;    /* warm yellow */
  --bg:#F7F9FC;        /* very light grey-blue */
  --text:#0F172A;
  --muted:#64748B;
  --card:#FFFFFF;
  --shadow: 0 6px 18px rgba(2,18,63,0.06);
  --radius:16px;
}
html, body, .block-container { background: var(--bg); }
#MainMenu, footer {visibility: hidden;}
.app-title {font-size:40px; font-weight:900; color:var(--text); text-align:center; letter-spacing:.2px;}
.app-subtitle {font-size:16px; color:var(--muted); text-align:center; margin-top:4px;}
.underline {width:140px; height:5px; background:var(--accent); margin:10px auto 26px auto; border-radius:999px;}

.card {background:var(--card); border:1px solid #eef2f7; border-radius:var(--radius); box-shadow:var(--shadow); padding:18px 18px;}
.card h3 {margin:0 0 8px 0; color:var(--text);}
.section {margin-top:18px;}

.pill {display:inline-block; background:#E5EDFF; color:#0B3D91; border:1px solid #C9D8FF; padding:6px 10px; border-radius:999px; font-weight:700; font-size:12px; margin-right:8px;}

.stButton > button, .stDownloadButton > button {
  background:var(--primary); color:#fff; border-radius:12px; padding:10px 16px; font-weight:700;
  border:2px solid var(--primary); box-shadow:var(--shadow);
}
.stButton > button:hover, .stDownloadButton > button:hover { background:#082e70; border-color:#082e70; }

hr {border:none; height:1px; background:#E9EEF5; margin:18px 0 10px 0;}
.small {color:var(--muted); font-size:12px;}
</style>
""", unsafe_allow_html=True)

st.markdown("<div class='app-title'>FLN Monitoring Report Summary</div>", unsafe_allow_html=True)
st.markdown("<div class='app-subtitle'>Generate FLN Block &amp; Cluster Monitoring Visit summary</div>", unsafe_allow_html=True)
st.markdown("<div class='underline'></div>", unsafe_allow_html=True)

# =====================
# Helpers
# =====================
def detect_columns(df: pd.DataFrame):
    cols = list(df.columns)

    def guess(patterns):
        for p in patterns:
            for c in cols:
                try:
                    if pd.Series([str(c)]).str.contains(p, case=False, regex=True).iloc[0]:
                        return c
                except Exception:
                    pass
        return None

    district = guess([r"district", r"जिला"])
    block    = guess([r"block", r"ब्लॉक|खंड"])
    cluster  = guess([r"cluster|sankul|jsk|jan\\s*shiksha", r"संकुल|जन शिक्ष"])
    role     = guess([r"role", r"पद"])
    target   = guess([r"target|लक्ष्य"])
    done     = guess([r"completed|done|स्थिति|visit.*done|actual|completion"])
    person   = guess([r"(mentor|visitor|employee|user|staff|teacher).*name", r"^name$", r"नाम", r"visitor", r"mentor name", r"employee name", r"user name"])

    return district, block, cluster, role, target, done, person

def block_summary_with_total(df, district_col, block_col, target_col, done_col, district_value):
    dff = df[df[district_col].astype(str).str.contains(str(district_value), case=False, na=False)].copy()
    agg = dff.groupby(block_col)[[target_col, done_col]].sum().reset_index()
    agg.rename(columns={block_col:"Block Name", target_col:"Targeted Visits", done_col:"Completed"}, inplace=True)
    agg["Monitoring %"] = np.where(agg["Targeted Visits"]>0, (agg["Completed"]/agg["Targeted Visits"])*100, np.nan).round(0)
    agg = agg.sort_values("Monitoring %", ascending=False)
    grand = pd.DataFrame({
        "Block Name":["Grand Total"],
        "Targeted Visits":[agg["Targeted Visits"].sum()],
        "Completed":[agg["Completed"].sum()],
        "Monitoring %":[round((agg["Completed"].sum()/agg["Targeted Visits"].sum())*100, 0) if agg["Targeted Visits"].sum() else np.nan]
    })
    return pd.concat([agg, grand], ignore_index=True)

def cluster_summary_with_total(df, district_col, block_col, cluster_col, target_col, done_col, district_value, block_value):
    dff = df[
        df[district_col].astype(str).str.contains(str(district_value), case=False, na=False) &
        (df[block_col].astype(str) == str(block_value))
    ].copy()
    agg = dff.groupby(cluster_col)[[target_col, done_col]].sum().reset_index()
    agg.rename(columns={cluster_col:"Cluster Name", target_col:"Targeted Visits", done_col:"Completed"}, inplace=True)
    agg["Monitoring %"] = np.where(agg["Targeted Visits"]>0, (agg["Completed"]/agg["Targeted Visits"])*100, np.nan).round(0)
    agg = agg.sort_values("Monitoring %", ascending=False)
    grand = pd.DataFrame({
        "Cluster Name":["Grand Total"],
        "Targeted Visits":[agg["Targeted Visits"].sum()],
        "Completed":[agg["Completed"].sum()],
        "Monitoring %":[round((agg["Completed"].sum()/agg["Targeted Visits"].sum())*100, 0) if agg["Targeted Visits"].sum() else np.nan]
    })
    return pd.concat([agg, grand], ignore_index=True)

def role_leaderboard_with_units(df, role_col, person_col, target_col, done_col, district_col, district_value, include_roles, unit_cols, top_n=10, bottom_n=10, extra_filters=None):
    base = df[df[district_col].astype(str).str.contains(str(district_value), case=False, na=False)].copy()
    if extra_filters:
        for col, val in extra_filters.items():
            base = base[base[col].astype(str) == str(val)]
    if role_col and include_roles:
        base = base[base[role_col].astype(str).str.upper().isin([r.upper() for r in include_roles])]
    if not person_col or person_col not in base.columns:
        return None, None
    group_keys = [person_col] + [c for c in unit_cols if c]
    grp = base.groupby(group_keys)[[target_col, done_col]].sum().reset_index()
    grp = grp.rename(columns={person_col:"Name"})
    grp["Monitoring %"] = np.where(grp[target_col]>0, (grp[done_col]/grp[target_col])*100, np.nan).round(0)
    grp = grp.rename(columns={target_col:"Targeted Visits", done_col:"Completed"})
    grp_f = grp[grp["Targeted Visits"]>0].copy()
    top = grp_f.sort_values(["Monitoring %","Completed"], ascending=[False, False]).head(top_n).reset_index(drop=True)
    bottom = grp_f.sort_values(["Monitoring %","Completed"], ascending=[True, True]).head(bottom_n).reset_index(drop=True)
    return top, bottom

def cac_leaderboard_all_in_block(df, district_col, block_col, cluster_col, role_col, person_col, target_col, done_col, district_value, block_value):
    base = df[
        (df[district_col].astype(str).str.contains(str(district_value), case=False, na=False)) &
        (df[block_col].astype(str) == str(block_value))
    ].copy()
    if role_col:
        base = base[base[role_col].astype(str).str.upper().str.contains("CAC")]
    if not person_col or person_col not in base.columns:
        return None
    grp = base.groupby([person_col, block_col, cluster_col])[[target_col, done_col]].sum().reset_index()
    grp = grp.rename(columns={person_col:"Name", block_col:"Block Name", cluster_col:"Cluster Name",
                              target_col:"Targeted Visits", done_col:"Completed"})
    grp["Monitoring %"] = np.where(grp["Targeted Visits"]>0, (grp["Completed"]/grp["Targeted Visits"])*100, np.nan).round(0)
    grp = grp.sort_values(["Monitoring %","Completed"], ascending=[False, False]).reset_index(drop=True)
    grp.insert(0, "Rank", np.arange(1, len(grp)+1))
    return grp

def style_top_bottom(df: pd.DataFrame):
    n = len(df.index)
    styles = pd.DataFrame('', index=df.index, columns=df.columns)
    if n > 0:
        top_n = min(3, n)
        bot_n = min(3, n)
        styles.iloc[:top_n, :] = 'background-color: #E8F5E9'  # soft green
        styles.iloc[n-bot_n:, :] = 'background-color: #FDECEC'  # soft red
    return styles

def excel_with_leaderboards(summary_df, leaderboards: dict, filename_prefix, unit_name, sheet_title="Summary"):
    with pd.ExcelWriter(io.BytesIO(), engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name=sheet_title, index=False)
        ws = writer.book[sheet_title]
        widths = [28, 18, 18, 16]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+i)].width = w
        header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
            for cell in row:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center")
        for r in range(2, ws.max_row+1):
            cell = ws.cell(row=r, column=4)
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value/100.0
        last = ws.max_row
        for cell in ws[last]:
            cell.font = Font(bold=True)
        rule = ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                              mid_type='num', mid_value=0.65, mid_color='FFEB84',
                              end_type='num', end_value=1, end_color='63BE7B')
        ws.conditional_formatting.add(f"D2:D{last}", rule)
        for sheet_name, df in leaderboards.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        writer.close()
        data = writer._handles.handle.getvalue()
    today = dt.datetime.now().strftime("%d-%m-%Y")
    filename = f"{filename_prefix}_{unit_name}_{today}.xlsx"
    return data, filename

def excel_with_sheets(summary_df, sheets: dict, filename_prefix, unit_name, sheet_title="Summary"):
    with pd.ExcelWriter(io.BytesIO(), engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name=sheet_title, index=False)
        ws = writer.book[sheet_title]
        widths = [28, 18, 18, 16]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64+i)].width = w
        header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
            for cell in row:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center")
        for r in range(2, ws.max_row+1):
            cell = ws.cell(row=r, column=4)
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value/100.0
        last = ws.max_row
        for cell in ws[last]:
            cell.font = Font(bold=True)
        rule = ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                              mid_type='num', mid_value=0.65, mid_color='FFEB84',
                              end_type='num', end_value=1, end_color='63BE7B')
        ws.conditional_formatting.add(f"D2:D{last}", rule)
        for sheet_name, df in sheets.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        writer.close()
        data = writer._handles.handle.getvalue()
    today = dt.datetime.now().strftime("%d-%m-%Y")
    filename = f"{filename_prefix}_{unit_name}_{today}.xlsx"
    return data, filename

# =====================
# Inputs Card
# =====================
with st.container():
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("### Upload CSV")
    file = st.file_uploader("Upload the raw state export (CSV)", type=["csv"])
    st.markdown("</div>", unsafe_allow_html=True)

if not file:
    st.info("Upload a CSV to begin.")
    st.stop()

# =====================
# After CSV uploaded
# =====================
df = pd.read_csv(file, encoding_errors="ignore")
district_col, block_col, cluster_col, role_col, target_col, done_col, person_col = detect_columns(df)

# -------- Auto-detect report type from Role column --------
options = ["CAC (Cluster / CACs)", "BRC+BAC (Block Mentors)"]
detected = "CAC (Cluster / CACs)"
if role_col and role_col in df.columns:
    roles = df[role_col].astype(str).str.upper()
    has_cac = roles.str.contains(r'\bCAC\b', regex=True).any()
    has_brcbac = roles.str.contains(r'\bBRCC?\b|\bBAC\b', regex=True).any()
    if has_cac and not has_brcbac:
        detected = options[0]
    elif has_brcbac and not has_cac:
        detected = options[1]
    else:
        detected = options[0]  # mixed – default to CAC but allow override
st.markdown(f"<div class='card'><b>Report type (auto‑detected):</b> <span class='pill'>{detected}</span></div>", unsafe_allow_html=True)
report_type = st.radio("Override if needed", options, index=options.index(detected), horizontal=True)
report_key = "CAC" if report_type.startswith("CAC") else "BRCBAC"

# Detected columns chip card
with st.container():
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    chip = lambda label, val: f"<span class='pill'>{label} → <b>{val or '—'}</b></span>"
    st.markdown("#### Detected columns")
    st.markdown(" ".join([
        chip("District", district_col), chip("Block", block_col),
        chip("Cluster", cluster_col or "—"), chip("Target", target_col),
        chip("Completed", done_col), chip("Role", role_col or "—"),
        chip("Name", person_col or "—")
    ]), unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

# =====================
# Section 1: District Level (Block Compliance)
# =====================
with st.container():
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("## District Level: Block Compliance Report")
    st.caption("Generate block-wise compliance and leaderboards for one or more districts.")
    mode = st.radio("", ["Selected districts", "All districts"], horizontal=True, label_visibility="collapsed")
    districts = sorted(df[district_col].dropna().unique().tolist())
    default_sel = [d for d in districts if 'jabalpur' in str(d).lower()]
    if mode == "Selected districts":
        sel = st.multiselect("Choose district(s)", districts, default=default_sel or districts[:1])
    else:
        sel = districts
    build = st.button("⚙️ Build District Reports", type="primary")
    st.markdown("</div>", unsafe_allow_html=True)

outputs = []
if build:
    progress = st.progress(0, text="Starting...")
    for i, dist in enumerate(sel, start=1):
        summary = block_summary_with_total(df, district_col, block_col, target_col, done_col, dist)
        leaderboards = {}
        if report_key == "BRCBAC":
            top_brcc, bottom_brcc = role_leaderboard_with_units(df, role_col, person_col, target_col, done_col, district_col, dist,
                                                                include_roles=["BRCC","BRC"], unit_cols=[block_col], top_n=3, bottom_n=3)
            top_bac, bottom_bac = role_leaderboard_with_units(df, role_col, person_col, target_col, done_col, district_col, dist,
                                                              include_roles=["BAC"], unit_cols=[block_col], top_n=3, bottom_n=3)
            leaderboards.update({
                "Top 3 BRCCs": top_brcc, "Bottom 3 BRCCs": bottom_brcc,
                "Top 3 BACs": top_bac, "Bottom 3 BACs": bottom_bac
            })
            prefix = "BRC_BAC_Monitoring"
        else:
            top_cac, bottom_cac = role_leaderboard_with_units(df, role_col, person_col, target_col, done_col, district_col, dist,
                                                              include_roles=["CAC"], unit_cols=[block_col, cluster_col], top_n=10, bottom_n=10)
            leaderboards.update({
                "Top 10 CACs": top_cac, "Bottom 10 CACs": bottom_cac
            })
            prefix = "CAC_Monitoring"
        excel_bytes, xlsx_name = excel_with_leaderboards(summary, leaderboards, prefix, dist, sheet_title="Block Summary")
        outputs.append((dist, summary, leaderboards, excel_bytes, xlsx_name))
        progress.progress(i/len(sel), text=f"Processed {i}/{len(sel)}: {dist}")
    progress.empty()

    if outputs:
        with st.container():
            st.markdown("<div class='card section'>", unsafe_allow_html=True)
            st.markdown("### Preview")
            st.dataframe(outputs[0][1], use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)

        with st.container():
            st.markdown("<div class='card section'>", unsafe_allow_html=True)
            st.markdown("### Leaderboards")
            dist, _, lbs, _, _ = outputs[0]
            st.caption(f"Showing for **{dist}**")
            for name, df_lb in lbs.items():
                if df_lb is not None and not df_lb.empty:
                    st.markdown(f"**{name}**")
                    st.dataframe(df_lb, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)

        with st.container():
            st.markdown("<div class='card section'>", unsafe_allow_html=True)
            st.markdown("### Downloads")
            for dist, _, _, excel_bytes, xlsx_name in outputs:
                st.download_button(
                    label=f"⬇️ Download Excel – {dist}",
                    data=excel_bytes,
                    file_name=xlsx_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            if len(outputs) > 1:
                mem_zip = io.BytesIO()
                with ZipFile(mem_zip, mode="w", compression=ZIP_DEFLATED) as zf:
                    for _, _, _, excel_bytes, xlsx_name in outputs:
                        zf.writestr(xlsx_name, excel_bytes)
                st.download_button(
                    label="⬇️ Download ALL as ZIP",
                    data=mem_zip.getvalue(),
                    file_name=f"FLN_{report_key}_Reports_{dt.datetime.now().strftime('%d-%m-%Y')}.zip",
                    mime="application/zip"
                )
            st.markdown("</div>", unsafe_allow_html=True)

        st.success(f"Generated {len(outputs)} district report(s).")

# =====================
# Section 2: Block Level (Cluster Compliance + CAC single leaderboard)
# =====================
with st.container():
    st.markdown("<div class='card'>", unsafe_allow_html=True)
    st.markdown("## Block Level: Cluster Compliance Report")
    st.caption("Pick district and block to view cluster-wise compliance, with a single CAC leaderboard (all CACs).")

    districts = sorted(df[district_col].dropna().unique().tolist())
    default_d = 0
    for i, d in enumerate(districts):
        if 'jabalpur' in str(d).lower():
            default_d = i; break
    d_sel = st.selectbox("Choose district", districts, index=default_d, key="blk_district_v11")

    blocks = sorted(df.loc[df[district_col].astype(str)==str(d_sel), block_col].dropna().unique().tolist()) if d_sel is not None else []
    b_sel = st.selectbox("Choose block", blocks, index=0 if blocks else None, key="blk_block_v11")

    go_blk = st.button("📘 Build Cluster Report", use_container_width=False, key="build_cluster_v11")

    if go_blk and b_sel and cluster_col:
        clus_sum = cluster_summary_with_total(df, district_col, block_col, cluster_col, target_col, done_col, d_sel, b_sel)
        st.markdown("#### Cluster Summary")
        st.dataframe(clus_sum, use_container_width=True)

        st.markdown("#### CAC Leaderboard (all CACs in this block)")
        cac_all = cac_leaderboard_all_in_block(
            df, district_col, block_col, cluster_col, role_col, person_col, target_col, done_col, d_sel, b_sel
        )
        if cac_all is not None and not cac_all.empty:
            styler = cac_all.style.apply(style_top_bottom, axis=None)
            st.dataframe(styler, use_container_width=True)
        else:
            st.info("No CAC Name column detected.")

        sheets = {"CAC Leaderboard": cac_all}
        excel_bytes, xlsx_name = excel_with_sheets(clus_sum, sheets, "Cluster_Monitoring", f"{d_sel}_{b_sel}", sheet_title="Cluster Summary")
        st.download_button("⬇️ Download Excel – Cluster view", data=excel_bytes, file_name=xlsx_name,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    elif go_blk and not cluster_col:
        st.warning("Cluster column not detected; cannot build cluster summary.")
    st.markdown("</div>", unsafe_allow_html=True)
